"""
PDF Table Extraction Dashboard
Multi-method extraction with fallback handling
"""

import os
import json
import io
import re
import csv
import traceback
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template_string, send_from_directory

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['EXPORT_FOLDER'] = 'exports'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)


def clean_cell(val):
    """
    Clean cell value with watermark artifact removal.
    BPS PDFs embed 'bps.go.id' watermark text that bleeds into table cells.
    Strategy:
      1. Remove known watermark patterns
      2. If cell looks numeric, strip stray letters
      3. Normalize dashes, whitespace
    """
    if val is None:
        return ''
    val = str(val).strip()
    val = val.replace('\n', ' ')

    # --- Step 1: Remove known BPS watermark fragments ---
    # Watermark chars that appear as isolated letters near numbers
    watermark_patterns = [
        r'\bbps\b', r'\bgo\b', r'\.id\b',
        r'\bwww\b', r'\bhttp\b',
        # Isolated single/double letters between spaces in otherwise numeric cells
    ]
    for pat in watermark_patterns:
        val = re.sub(pat, '', val, flags=re.IGNORECASE)

    # --- Step 2: Collapse multiple spaces ---
    val = re.sub(r'\s+', ' ', val).strip()

    # --- Step 3: Detect if cell is NUMERIC type ---
    # Indonesian number format: 1.234,56 or –
    # Strip stray letters if cell contains digits + Indonesian number punctuation
    has_digits = bool(re.search(r'\d', val))
    has_dash = val.strip() in ('–', '-', '−', '--')

    if has_digits:
        # Check if it's clearly a corrupted number:
        # Contains digits and commas/dots but also has stray letters
        # Pattern: optional letters/spaces before/between number chars
        stripped = re.sub(r'[^\d.,–\-−]', '', val)
        # Rebuild: keep the cleaned numeric form if it still has digits
        if re.search(r'\d', stripped):
            val = stripped

    # --- Step 4: Normalize dash variants ---
    # Clean dash-only cells
    clean = re.sub(r'\s', '', val)
    if clean in ('', '-', '–', '−', '--', '–-', '-–'):
        return '–'

    # --- Step 5: Fix common OCR/watermark digit corruptions ---
    # Remove stray leading/trailing single letters after cleaning
    val = re.sub(r'^[a-zA-Z]\s+', '', val)  # leading stray letter
    val = re.sub(r'\s+[a-zA-Z]$', '', val)  # trailing stray letter
    val = val.strip()

    return val


def clean_table(table_data):
    """Post-process full table: detect column types and clean accordingly."""
    if not table_data:
        return table_data

    # Determine which columns are likely numeric (>50% of non-header rows are numbers)
    num_cols = len(table_data[0]) if table_data else 0
    col_numeric = [False] * num_cols

    data_rows = table_data[1:] if len(table_data) > 1 else table_data
    for col_i in range(num_cols):
        vals = [r[col_i] for r in data_rows if col_i < len(r)]
        numeric_count = 0
        for v in vals:
            cleaned = re.sub(r'[^\d]', '', v)
            if cleaned or v in ('–', '-', '−', ''):
                numeric_count += 1
        if vals and numeric_count / len(vals) > 0.5:
            col_numeric[col_i] = True

    # Second pass: for numeric columns, aggressively strip non-numeric chars
    result = []
    for ri, row in enumerate(table_data):
        new_row = []
        for ci, cell in enumerate(row):
            if ri > 0 and ci < num_cols and col_numeric[ci]:
                # Aggressively clean: keep only digits, . , – -
                cleaned = re.sub(r'[^\d.,\-–−]', '', cell)
                if not cleaned:
                    cleaned = '–' if '–' in cell or '-' in cell else ''
                new_row.append(cleaned)
            else:
                new_row.append(cell)
        result.append(new_row)

    return result


def extract_tables_pdfplumber(pdf_path, page_num):
    """Primary method: pdfplumber with watermark filtering + tuned settings."""
    import pdfplumber, statistics
    results = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if page_num < 1 or page_num > len(pdf.pages):
                return results, f"Halaman {page_num} tidak ada (total {len(pdf.pages)} halaman)"
            
            page = pdf.pages[page_num - 1]

            # --- Filter watermark by font size ---
            # BPS watermarks are large, semi-transparent chars.
            # Keep only chars whose size is within 50% of the median.
            try:
                chars = page.chars
                sizes = [c['size'] for c in chars if c.get('size')]
                if sizes:
                    med = statistics.median(sizes)
                    page = page.filter(
                        lambda obj: obj.get('object_type') != 'char' or
                                    abs(obj.get('size', med) - med) <= med * 0.5
                    )
            except Exception:
                pass

            # Try default settings first
            tables = page.extract_tables()
            
            if not tables:
                ts = {
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "intersection_tolerance": 10,
                }
                tables = page.extract_tables(table_settings=ts)
            
            if not tables:
                ts2 = {
                    "vertical_strategy": "text",
                    "horizontal_strategy": "text",
                    "snap_tolerance": 5,
                }
                tables = page.extract_tables(table_settings=ts2)
            
            for i, table in enumerate(tables):
                if not table:
                    continue
                # Basic clean
                cleaned = [[clean_cell(c) for c in row] for row in table if any(c for c in row)]
                # Type-aware watermark removal pass
                cleaned = clean_table(cleaned)
                if cleaned:
                    results.append({
                        'index': i,
                        'rows': len(cleaned),
                        'cols': max(len(r) for r in cleaned),
                        'data': cleaned,
                        'method': 'pdfplumber'
                    })
        return results, None
    except Exception as e:
        return results, str(e)


def extract_tables_camelot(pdf_path, page_num):
    """Fallback method: camelot for complex tables."""
    results = []
    try:
        import camelot
        tables = camelot.read_pdf(str(pdf_path), pages=str(page_num), flavor='lattice')
        
        if len(tables) == 0:
            tables = camelot.read_pdf(str(pdf_path), pages=str(page_num), flavor='stream')
        
        for i, table in enumerate(tables):
            df = table.df
            data = [df.columns.tolist()] + df.values.tolist()
            cleaned = [[clean_cell(c) for c in row] for row in data]
            results.append({
                'index': i,
                'rows': len(cleaned),
                'cols': len(cleaned[0]) if cleaned else 0,
                'data': cleaned,
                'method': 'camelot',
                'accuracy': round(table.accuracy, 1)
            })
        return results, None
    except Exception as e:
        return results, str(e)


def extract_text_page(pdf_path, page_num):
    """Extract raw text from a page for keyword search."""
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            if page_num < 1 or page_num > len(pdf.pages):
                return ''
            page = pdf.pages[page_num - 1]
            return page.extract_text() or ''
    except:
        return ''


def search_keyword_in_pdf(pdf_path, keyword):
    """Search keyword across all pages, return matching page numbers."""
    import pdfplumber
    matches = []
    keyword_lower = keyword.lower()
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ''
                if keyword_lower in text.lower():
                    # Get context (first matching line)
                    lines = text.split('\n')
                    context_lines = [l.strip() for l in lines if keyword_lower in l.lower()]
                    context = context_lines[0][:100] if context_lines else ''
                    matches.append({
                        'page': i + 1,
                        'context': context
                    })
        return matches, total
    except Exception as e:
        return [], 0


def get_pdf_info(pdf_path):
    """Get basic PDF info."""
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            return {
                'pages': len(pdf.pages),
                'filename': Path(pdf_path).name
            }
    except:
        return {'pages': 0, 'filename': Path(pdf_path).name}


# ─── API ROUTES ───────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/api/upload', methods=['POST'])
def upload_pdf():
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file'}), 400
    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'File harus berformat PDF'}), 400
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)
    info = get_pdf_info(filepath)
    return jsonify({'success': True, 'filename': file.filename, **info})


@app.route('/api/files')
def list_files():
    files = []
    folder = app.config['UPLOAD_FOLDER']
    for f in os.listdir(folder):
        if f.lower().endswith('.pdf'):
            fp = os.path.join(folder, f)
            info = get_pdf_info(fp)
            files.append({'filename': f, 'pages': info['pages'],
                          'size': round(os.path.getsize(fp) / 1024 / 1024, 2)})
    return jsonify(files)


@app.route('/api/pdf/<filename>')
def serve_pdf(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/api/search', methods=['POST'])
def search():
    data = request.json
    filename = data.get('filename')
    keyword = data.get('keyword', '').strip()
    
    if not filename or not keyword:
        return jsonify({'error': 'Filename dan keyword diperlukan'}), 400
    
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(pdf_path):
        return jsonify({'error': 'File tidak ditemukan'}), 404
    
    matches, total = search_keyword_in_pdf(pdf_path, keyword)
    return jsonify({'matches': matches, 'total_pages': total, 'keyword': keyword})


@app.route('/api/extract', methods=['POST'])
def extract():
    data = request.json
    filename = data.get('filename')
    page_num = int(data.get('page', 1))
    method = data.get('method', 'auto')  # auto, pdfplumber, camelot
    
    if not filename:
        return jsonify({'error': 'Filename diperlukan'}), 400
    
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(pdf_path):
        return jsonify({'error': 'File tidak ditemukan'}), 404
    
    errors = []
    tables = []
    
    if method in ('auto', 'pdfplumber'):
        tables, err = extract_tables_pdfplumber(pdf_path, page_num)
        if err:
            errors.append(f'pdfplumber: {err}')
    
    # If pdfplumber found nothing or method is camelot, try camelot
    if (not tables and method == 'auto') or method == 'camelot':
        tables_c, err = extract_tables_camelot(pdf_path, page_num)
        if err:
            errors.append(f'camelot: {err}')
        if tables_c:
            tables = tables_c
    
    # Get page text for context
    page_text = extract_text_page(pdf_path, page_num)
    
    return jsonify({
        'tables': tables,
        'page': page_num,
        'page_text_preview': page_text[:300] if page_text else '',
        'errors': errors,
        'found': len(tables)
    })


@app.route('/api/export', methods=['POST'])
def export_table():
    data = request.json
    table_data = data.get('data', [])
    fmt = data.get('format', 'csv')  # csv, json, excel
    filename = data.get('filename', 'tabel_ekstrak')
    
    if not table_data:
        return jsonify({'error': 'Data tabel kosong'}), 400
    
    safe_name = re.sub(r'[^\w\-]', '_', filename)
    export_path = os.path.join(app.config['EXPORT_FOLDER'], safe_name)
    
    if fmt == 'csv':
        out_path = f'{export_path}.csv'
        with open(out_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerows(table_data)
        return send_file(out_path, as_attachment=True,
                         download_name=f'{safe_name}.csv',
                         mimetype='text/csv')
    
    elif fmt == 'json':
        if len(table_data) > 1:
            headers = table_data[0]
            rows = []
            for row in table_data[1:]:
                obj = {}
                for j, h in enumerate(headers):
                    obj[h or f'col_{j}'] = row[j] if j < len(row) else ''
                rows.append(obj)
            output = rows
        else:
            output = table_data
        
        out_path = f'{export_path}.json'
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        return send_file(out_path, as_attachment=True,
                         download_name=f'{safe_name}.json',
                         mimetype='application/json')
    
    elif fmt == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Data Ekstrak'
            
            header_fill = PatternFill(start_color='1E4D8C', end_color='1E4D8C', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            alt_fill = PatternFill(start_color='EEF2F8', end_color='EEF2F8', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for ri, row in enumerate(table_data):
                for ci, val in enumerate(row):
                    cell = ws.cell(row=ri+1, column=ci+1, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    if ri == 0:
                        cell.fill = header_fill
                        cell.font = header_font
                    elif ri % 2 == 0:
                        cell.fill = alt_fill
            
            # Auto column width
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
            
            out_path = f'{export_path}.xlsx'
            wb.save(out_path)
            return send_file(out_path, as_attachment=True,
                             download_name=f'{safe_name}.xlsx',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            return jsonify({'error': f'Export Excel gagal: {str(e)}'}), 500
    
    return jsonify({'error': 'Format tidak dikenal'}), 400


# ─── HTML TEMPLATE ─────────────────────────────────────────────────────────────

HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PDF Table Extractor Dashboard</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=Sora:wght@300;400;600;700&display=swap');

  :root {
    --bg: #0d1117;
    --surface: #161b22;
    --surface2: #1c2230;
    --border: #2a3447;
    --accent: #3b82f6;
    --accent2: #10b981;
    --accent3: #f59e0b;
    --danger: #ef4444;
    --text: #e6edf3;
    --muted: #7d8fa8;
    --mono: 'IBM Plex Mono', monospace;
    --sans: 'Sora', sans-serif;
  }

  * { box-sizing: border-box; margin: 0; padding: 0; }
  
  body {
    background: var(--bg);
    color: var(--text);
    font-family: var(--sans);
    min-height: 100vh;
    display: flex;
    flex-direction: column;
  }

  /* Header */
  header {
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    padding: 0 24px;
    height: 56px;
    display: flex;
    align-items: center;
    gap: 16px;
    position: sticky;
    top: 0;
    z-index: 100;
  }
  .logo {
    font-family: var(--mono);
    font-size: 14px;
    font-weight: 600;
    color: var(--accent);
    letter-spacing: -0.5px;
    display: flex;
    align-items: center;
    gap: 8px;
  }
  .logo-icon {
    width: 28px; height: 28px;
    background: var(--accent);
    border-radius: 6px;
    display: flex; align-items: center; justify-content: center;
    font-size: 14px;
  }
  .header-status {
    margin-left: auto;
    font-size: 12px;
    color: var(--muted);
    font-family: var(--mono);
  }

  /* Layout */
  .main {
    display: grid;
    grid-template-columns: 300px 1fr;
    flex: 1;
    min-height: 0;
  }

  /* Sidebar */
  .sidebar {
    background: var(--surface);
    border-right: 1px solid var(--border);
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }
  .sidebar-section {
    padding: 16px;
    border-bottom: 1px solid var(--border);
  }
  .sidebar-label {
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 12px;
  }

  /* Upload zone */
  .upload-zone {
    border: 2px dashed var(--border);
    border-radius: 10px;
    padding: 20px;
    text-align: center;
    cursor: pointer;
    transition: all 0.2s;
  }
  .upload-zone:hover, .upload-zone.drag {
    border-color: var(--accent);
    background: rgba(59,130,246,0.05);
  }
  .upload-zone .up-icon { font-size: 28px; margin-bottom: 8px; }
  .upload-zone .up-text { font-size: 12px; color: var(--muted); }
  .upload-zone .up-link { color: var(--accent); cursor: pointer; }
  #fileInput { display: none; }

  /* File list */
  .file-list { overflow-y: auto; max-height: 180px; }
  .file-item {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 10px 12px;
    border-radius: 8px;
    cursor: pointer;
    transition: background 0.15s;
    border: 1px solid transparent;
  }
  .file-item:hover { background: var(--surface2); }
  .file-item.active {
    background: rgba(59,130,246,0.1);
    border-color: rgba(59,130,246,0.3);
  }
  .file-icon { font-size: 18px; flex-shrink: 0; }
  .file-info { flex: 1; min-width: 0; }
  .file-name { font-size: 12px; font-weight: 600; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .file-meta { font-size: 10px; color: var(--muted); font-family: var(--mono); }

  /* Controls */
  .ctrl-row { display: flex; gap: 8px; margin-bottom: 10px; align-items: center; }
  label.ctrl-label { font-size: 11px; color: var(--muted); display: block; margin-bottom: 4px; }
  
  input[type=text], input[type=number], select {
    background: var(--surface2);
    border: 1px solid var(--border);
    color: var(--text);
    border-radius: 6px;
    padding: 8px 10px;
    font-size: 13px;
    font-family: var(--sans);
    width: 100%;
    outline: none;
    transition: border-color 0.15s;
  }
  input:focus, select:focus { border-color: var(--accent); }

  .btn {
    border: none;
    border-radius: 6px;
    padding: 8px 14px;
    font-size: 12px;
    font-weight: 600;
    font-family: var(--sans);
    cursor: pointer;
    transition: all 0.15s;
    white-space: nowrap;
    display: inline-flex; align-items: center; gap: 6px;
  }
  .btn-primary { background: var(--accent); color: white; }
  .btn-primary:hover { background: #2563eb; }
  .btn-success { background: var(--accent2); color: white; }
  .btn-success:hover { background: #059669; }
  .btn-warn { background: var(--accent3); color: #1a1a1a; }
  .btn-warn:hover { background: #d97706; }
  .btn-ghost {
    background: transparent;
    color: var(--muted);
    border: 1px solid var(--border);
  }
  .btn-ghost:hover { border-color: var(--accent); color: var(--accent); }
  .btn-sm { padding: 5px 10px; font-size: 11px; }
  .btn:disabled { opacity: 0.4; cursor: not-allowed; }

  /* Search results */
  .search-hits { overflow-y: auto; max-height: 200px; }
  .hit-item {
    padding: 8px 10px;
    border-radius: 6px;
    cursor: pointer;
    transition: background 0.15s;
    border-bottom: 1px solid var(--border);
  }
  .hit-item:hover { background: var(--surface2); }
  .hit-page {
    font-family: var(--mono);
    font-size: 11px;
    color: var(--accent);
    font-weight: 600;
  }
  .hit-ctx { font-size: 11px; color: var(--muted); margin-top: 2px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }

  /* Main content */
  .content {
    display: grid;
    grid-template-rows: auto 1fr;
    overflow: hidden;
  }

  /* Tab bar */
  .tab-bar {
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    display: flex;
    padding: 0 20px;
    gap: 4px;
  }
  .tab {
    padding: 14px 18px;
    font-size: 13px;
    font-weight: 500;
    color: var(--muted);
    cursor: pointer;
    border-bottom: 2px solid transparent;
    transition: all 0.15s;
    display: flex; align-items: center; gap: 6px;
  }
  .tab:hover { color: var(--text); }
  .tab.active { color: var(--accent); border-bottom-color: var(--accent); }

  /* Panels */
  .panels { overflow: hidden; position: relative; }
  .panel { display: none; height: 100%; overflow: auto; padding: 20px; }
  .panel.active { display: block; }

  /* PDF Preview */
  #pdfFrame {
    width: 100%;
    height: calc(100vh - 160px);
    border: none;
    border-radius: 8px;
    background: #fff;
  }
  .no-preview {
    height: 300px;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    color: var(--muted);
    border: 2px dashed var(--border);
    border-radius: 12px;
    gap: 12px;
  }
  .no-preview .np-icon { font-size: 48px; }

  /* Table panel */
  .extract-header {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 16px;
    flex-wrap: wrap;
  }
  .page-badge {
    font-family: var(--mono);
    font-size: 12px;
    background: rgba(59,130,246,0.15);
    color: var(--accent);
    padding: 4px 10px;
    border-radius: 20px;
    border: 1px solid rgba(59,130,246,0.3);
  }
  .method-badge {
    font-family: var(--mono);
    font-size: 11px;
    padding: 3px 8px;
    border-radius: 10px;
  }
  .method-pdfplumber { background: rgba(16,185,129,0.15); color: var(--accent2); border: 1px solid rgba(16,185,129,0.3); }
  .method-camelot { background: rgba(245,158,11,0.15); color: var(--accent3); border: 1px solid rgba(245,158,11,0.3); }

  /* Table container */
  .table-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 10px;
    margin-bottom: 20px;
    overflow: hidden;
  }
  .table-card-header {
    padding: 12px 16px;
    display: flex;
    align-items: center;
    gap: 10px;
    border-bottom: 1px solid var(--border);
    background: var(--surface2);
  }
  .table-title { font-size: 13px; font-weight: 600; }
  .table-meta { font-size: 11px; color: var(--muted); font-family: var(--mono); margin-left: auto; }
  .table-wrapper { overflow-x: auto; }
  
  table.data-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 12px;
  }
  table.data-table th {
    background: rgba(59,130,246,0.12);
    color: var(--accent);
    font-weight: 600;
    padding: 10px 12px;
    text-align: left;
    border-bottom: 1px solid var(--border);
    white-space: nowrap;
    font-family: var(--mono);
    font-size: 11px;
  }
  table.data-table td {
    padding: 9px 12px;
    border-bottom: 1px solid rgba(42,52,71,0.5);
    color: var(--text);
    max-width: 200px;
    word-break: break-word;
  }
  table.data-table tr:last-child td { border-bottom: none; }
  table.data-table tr:hover td { background: rgba(255,255,255,0.02); }

  /* Export bar */
  .export-bar {
    display: flex;
    gap: 8px;
    align-items: center;
    padding: 12px 16px;
    background: var(--surface2);
    border-top: 1px solid var(--border);
    flex-wrap: wrap;
  }
  .export-label { font-size: 12px; color: var(--muted); font-weight: 500; }

  /* Status / error */
  .status-bar {
    padding: 10px 16px;
    border-radius: 8px;
    font-size: 12px;
    font-family: var(--mono);
    margin-bottom: 16px;
    display: none;
  }
  .status-bar.info { background: rgba(59,130,246,0.1); color: var(--accent); border: 1px solid rgba(59,130,246,0.2); display: block; }
  .status-bar.error { background: rgba(239,68,68,0.1); color: var(--danger); border: 1px solid rgba(239,68,68,0.2); display: block; }
  .status-bar.success { background: rgba(16,185,129,0.1); color: var(--accent2); border: 1px solid rgba(16,185,129,0.2); display: block; }

  /* Loading spinner */
  .spinner {
    width: 18px; height: 18px;
    border: 2px solid rgba(59,130,246,0.3);
    border-top-color: var(--accent);
    border-radius: 50%;
    animation: spin 0.7s linear infinite;
    display: inline-block;
  }
  @keyframes spin { to { transform: rotate(360deg); } }

  .empty-state {
    text-align: center;
    padding: 60px 20px;
    color: var(--muted);
  }
  .empty-state .es-icon { font-size: 48px; margin-bottom: 12px; }
  .empty-state .es-title { font-size: 16px; font-weight: 600; color: var(--text); margin-bottom: 6px; }
  .empty-state .es-desc { font-size: 13px; }

  /* Stats row */
  .stats-row {
    display: flex;
    gap: 12px;
    margin-bottom: 16px;
  }
  .stat-card {
    flex: 1;
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 12px 16px;
  }
  .stat-val { font-size: 24px; font-weight: 700; font-family: var(--mono); color: var(--accent); }
  .stat-label { font-size: 11px; color: var(--muted); margin-top: 2px; }

  .tooltip { position: relative; }
  .tooltip-text {
    display: none;
    position: absolute;
    bottom: 100%;
    left: 50%;
    transform: translateX(-50%);
    background: #333;
    color: white;
    font-size: 11px;
    padding: 4px 8px;
    border-radius: 4px;
    white-space: nowrap;
    margin-bottom: 4px;
    z-index: 10;
  }
  .tooltip:hover .tooltip-text { display: block; }

  /* Scrollbars */
  ::-webkit-scrollbar { width: 6px; height: 6px; }
  ::-webkit-scrollbar-track { background: transparent; }
  ::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
  ::-webkit-scrollbar-thumb:hover { background: var(--muted); }
</style>
</head>
<body>

<header>
  <div class="logo">
    <div class="logo-icon">📄</div>
    PDF Table Extractor
  </div>
  <span class="header-status" id="headerStatus">Belum ada file dipilih</span>
</header>

<div class="main">
  <!-- Sidebar -->
  <div class="sidebar">
    
    <!-- Upload -->
    <div class="sidebar-section">
      <div class="sidebar-label">📁 File PDF</div>
      <div class="upload-zone" id="uploadZone" onclick="document.getElementById('fileInput').click()">
        <div class="up-icon">⬆️</div>
        <div class="up-text">Drag & drop atau <span class="up-link">pilih file PDF</span></div>
      </div>
      <input type="file" id="fileInput" accept=".pdf" multiple>
    </div>

    <!-- File list -->
    <div class="sidebar-section" style="flex:0 0 auto">
      <div class="sidebar-label">📋 Daftar File</div>
      <div class="file-list" id="fileList">
        <div style="font-size:12px;color:var(--muted);text-align:center;padding:20px">Belum ada file</div>
      </div>
    </div>

    <!-- Search -->
    <div class="sidebar-section">
      <div class="sidebar-label">🔍 Cari Keyword</div>
      <div style="margin-bottom:8px">
        <input type="text" id="searchInput" placeholder="Contoh: Tabel 5.7, Pendidikan..." />
      </div>
      <button class="btn btn-primary" style="width:100%" onclick="doSearch()">
        🔍 Cari di Semua Halaman
      </button>
    </div>

    <!-- Search Results -->
    <div class="sidebar-section" style="flex:1; overflow:hidden; display:flex; flex-direction:column">
      <div class="sidebar-label">📌 Hasil Pencarian</div>
      <div class="search-hits" id="searchHits">
        <div style="font-size:12px;color:var(--muted);text-align:center;padding:20px">Belum ada pencarian</div>
      </div>
    </div>

    <!-- Extract control -->
    <div class="sidebar-section">
      <div class="sidebar-label">⚡ Ekstraksi Tabel</div>
      <div style="margin-bottom:10px">
        <label class="ctrl-label">Halaman</label>
        <input type="number" id="pageInput" value="1" min="1" />
      </div>
      <div style="margin-bottom:10px">
        <label class="ctrl-label">Metode Ekstraksi</label>
        <select id="methodSelect">
          <option value="auto">🤖 Auto (pdfplumber → camelot)</option>
          <option value="pdfplumber">📐 pdfplumber (garis tabel)</option>
          <option value="camelot">🗂️ camelot (lattice/stream)</option>
        </select>
      </div>
      <button class="btn btn-success" style="width:100%" id="extractBtn" onclick="doExtract()">
        ⚡ Ekstrak Tabel
      </button>
    </div>

  </div>

  <!-- Content -->
  <div class="content">
    <div class="tab-bar">
      <div class="tab active" onclick="switchTab('preview', this)">👁️ Preview PDF</div>
      <div class="tab" onclick="switchTab('extract', this)">📊 Hasil Ekstraksi</div>
    </div>

    <div class="panels">
      <!-- Preview panel -->
      <div class="panel active" id="panel-preview">
        <div class="no-preview" id="noPreview">
          <div class="np-icon">📄</div>
          <div style="font-size:16px; font-weight:600">Pilih file PDF</div>
          <div style="font-size:13px; color:var(--muted)">Upload atau pilih file dari daftar untuk preview</div>
        </div>
        <iframe id="pdfFrame" style="display:none"></iframe>
      </div>

      <!-- Extract panel -->
      <div class="panel" id="panel-extract">
        <div id="extractStatus" class="status-bar"></div>
        
        <div id="extractEmpty" class="empty-state">
          <div class="es-icon">📊</div>
          <div class="es-title">Siap Mengekstrak</div>
          <div class="es-desc">Pilih file, tentukan halaman, lalu klik "Ekstrak Tabel"</div>
        </div>

        <div id="extractResults" style="display:none">
          <div class="extract-header" id="extractHeader"></div>
          <div class="stats-row" id="statsRow"></div>
          <div id="tablesContainer"></div>
        </div>
      </div>
    </div>
  </div>
</div>

<script>
let currentFile = null;
let lastExtractedTables = [];

// ── Tab ──────────────────────────────────────────────────────────────────────
function switchTab(id, el) {
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
  el.classList.add('active');
  document.getElementById('panel-' + id).classList.add('active');
}

// ── Upload & File List ────────────────────────────────────────────────────────
const uploadZone = document.getElementById('uploadZone');
const fileInput = document.getElementById('fileInput');

uploadZone.addEventListener('dragover', e => { e.preventDefault(); uploadZone.classList.add('drag'); });
uploadZone.addEventListener('dragleave', () => uploadZone.classList.remove('drag'));
uploadZone.addEventListener('drop', e => {
  e.preventDefault(); uploadZone.classList.remove('drag');
  uploadFiles(e.dataTransfer.files);
});
fileInput.addEventListener('change', () => uploadFiles(fileInput.files));

async function uploadFiles(files) {
  for (const file of files) {
    if (!file.name.toLowerCase().endsWith('.pdf')) continue;
    const fd = new FormData();
    fd.append('file', file);
    try {
      showHeaderStatus('⬆️ Mengupload ' + file.name + '...');
      const r = await fetch('/api/upload', { method: 'POST', body: fd });
      const d = await r.json();
      if (d.success) {
        showHeaderStatus('✅ Upload berhasil: ' + file.name);
        loadFileList();
      } else {
        showHeaderStatus('❌ ' + (d.error || 'Upload gagal'));
      }
    } catch (e) {
      showHeaderStatus('❌ Error: ' + e.message);
    }
  }
}

async function loadFileList() {
  const r = await fetch('/api/files');
  const files = await r.json();
  const list = document.getElementById('fileList');
  
  if (!files.length) {
    list.innerHTML = '<div style="font-size:12px;color:var(--muted);text-align:center;padding:20px">Belum ada file</div>';
    return;
  }
  
  list.innerHTML = files.map(f => `
    <div class="file-item ${currentFile === f.filename ? 'active' : ''}" 
         onclick="selectFile('${f.filename}')">
      <div class="file-icon">📄</div>
      <div class="file-info">
        <div class="file-name" title="${f.filename}">${f.filename}</div>
        <div class="file-meta">${f.pages} hal · ${f.size} MB</div>
      </div>
    </div>
  `).join('');
}

function selectFile(filename) {
  currentFile = filename;
  document.getElementById('headerStatus').textContent = '📄 ' + filename;
  
  // Update active item
  document.querySelectorAll('.file-item').forEach(el => el.classList.remove('active'));
  event.currentTarget.classList.add('active');
  
  // Load preview
  const frame = document.getElementById('pdfFrame');
  const noPreview = document.getElementById('noPreview');
  frame.src = '/api/pdf/' + encodeURIComponent(filename);
  frame.style.display = 'block';
  noPreview.style.display = 'none';
}

function showHeaderStatus(msg) {
  document.getElementById('headerStatus').textContent = msg;
}

// ── Search ────────────────────────────────────────────────────────────────────
async function doSearch() {
  const keyword = document.getElementById('searchInput').value.trim();
  if (!keyword) return alert('Masukkan keyword pencarian');
  if (!currentFile) return alert('Pilih file PDF terlebih dahulu');

  const hitsDiv = document.getElementById('searchHits');
  hitsDiv.innerHTML = '<div style="text-align:center;padding:20px"><div class="spinner"></div></div>';

  const r = await fetch('/api/search', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({filename: currentFile, keyword})
  });
  const d = await r.json();

  if (d.error) {
    hitsDiv.innerHTML = `<div style="color:var(--danger);font-size:12px;padding:10px">${d.error}</div>`;
    return;
  }

  if (!d.matches.length) {
    hitsDiv.innerHTML = `<div style="color:var(--muted);font-size:12px;padding:10px;text-align:center">Tidak ditemukan "<b>${keyword}</b>"</div>`;
    return;
  }

  hitsDiv.innerHTML = d.matches.map(m => `
    <div class="hit-item" onclick="goToPage(${m.page})">
      <div class="hit-page">Hal. ${m.page}</div>
      <div class="hit-ctx">${escHtml(m.context)}</div>
    </div>
  `).join('');
}

function goToPage(page) {
  document.getElementById('pageInput').value = page;
  // Navigate PDF viewer to page
  const frame = document.getElementById('pdfFrame');
  if (frame.src) {
    frame.src = '/api/pdf/' + encodeURIComponent(currentFile) + '#page=' + page;
  }
  // Switch to preview tab
  document.querySelectorAll('.tab')[0].click();
}

// ── Extract ───────────────────────────────────────────────────────────────────
async function doExtract() {
  if (!currentFile) return alert('Pilih file PDF terlebih dahulu');
  
  const page = parseInt(document.getElementById('pageInput').value);
  const method = document.getElementById('methodSelect').value;
  
  const btn = document.getElementById('extractBtn');
  btn.disabled = true;
  btn.innerHTML = '<span class="spinner"></span> Mengekstrak...';
  
  const statusEl = document.getElementById('extractStatus');
  showStatus(statusEl, 'info', `⏳ Mengekstrak tabel dari halaman ${page} menggunakan metode ${method}...`);
  
  document.getElementById('extractEmpty').style.display = 'none';
  document.getElementById('extractResults').style.display = 'none';
  
  // Switch to extract tab
  document.querySelectorAll('.tab')[1].click();

  try {
    const r = await fetch('/api/extract', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({filename: currentFile, page, method})
    });
    const d = await r.json();
    
    btn.disabled = false;
    btn.innerHTML = '⚡ Ekstrak Tabel';
    
    if (d.error) {
      showStatus(statusEl, 'error', '❌ ' + d.error);
      document.getElementById('extractEmpty').style.display = 'block';
      return;
    }

    if (!d.tables.length) {
      showStatus(statusEl, 'error', `⚠️ Tidak ada tabel ditemukan di halaman ${page}. Coba metode lain atau periksa halaman yang benar.`);
      if (d.errors.length) {
        showStatus(statusEl, 'error', `⚠️ Tidak ada tabel di halaman ${page}. Error: ${d.errors.join('; ')}`);
      }
      document.getElementById('extractEmpty').style.display = 'block';
      return;
    }

    statusEl.style.display = 'none';
    lastExtractedTables = d.tables;
    renderExtractResults(d);

  } catch(e) {
    btn.disabled = false;
    btn.innerHTML = '⚡ Ekstrak Tabel';
    showStatus(statusEl, 'error', '❌ Error: ' + e.message);
  }
}

function renderExtractResults(d) {
  const results = document.getElementById('extractResults');
  results.style.display = 'block';

  // Header
  const header = document.getElementById('extractHeader');
  header.innerHTML = `
    <span class="page-badge">📄 Halaman ${d.page}</span>
    ${d.tables.map(t => `<span class="method-badge method-${t.method}">${t.method}${t.accuracy !== undefined ? ' · acc: ' + t.accuracy + '%' : ''}</span>`).join('')}
    ${d.errors.length ? `<span style="font-size:11px;color:var(--accent3)">⚠️ ${d.errors.join(' | ')}</span>` : ''}
  `;

  // Stats
  const totalRows = d.tables.reduce((s, t) => s + t.rows, 0);
  const totalCols = d.tables.length ? d.tables[0].cols : 0;
  document.getElementById('statsRow').innerHTML = `
    <div class="stat-card"><div class="stat-val">${d.tables.length}</div><div class="stat-label">Tabel Ditemukan</div></div>
    <div class="stat-card"><div class="stat-val">${totalRows}</div><div class="stat-label">Total Baris</div></div>
    <div class="stat-card"><div class="stat-val">${totalCols}</div><div class="stat-label">Kolom (tabel 1)</div></div>
  `;

  // Tables
  const container = document.getElementById('tablesContainer');
  container.innerHTML = d.tables.map((t, i) => `
    <div class="table-card">
      <div class="table-card-header">
        <span class="table-title">Tabel ${i + 1}</span>
        <span class="method-badge method-${t.method}">${t.method}</span>
        <span class="table-meta">${t.rows} baris × ${t.cols} kolom</span>
      </div>
      <div class="table-wrapper">
        ${renderTable(t.data)}
      </div>
      <div class="export-bar">
        <span class="export-label">Export tabel ini:</span>
        <button class="btn btn-primary btn-sm" onclick="exportTable(${i}, 'csv')">📥 CSV</button>
        <button class="btn btn-success btn-sm" onclick="exportTable(${i}, 'json')">📥 JSON</button>
        <button class="btn btn-warn btn-sm" onclick="exportTable(${i}, 'excel')">📥 Excel</button>
      </div>
    </div>
  `).join('');
}

function renderTable(data) {
  if (!data || !data.length) return '<div style="padding:20px;color:var(--muted)">Tidak ada data</div>';
  const [header, ...rows] = data;
  return `<table class="data-table">
    <thead><tr>${header.map(h => `<th>${escHtml(h)}</th>`).join('')}</tr></thead>
    <tbody>${rows.map(row => `<tr>${row.map(c => `<td>${escHtml(c)}</td>`).join('')}</tr>`).join('')}</tbody>
  </table>`;
}

async function exportTable(tableIdx, fmt) {
  const table = lastExtractedTables[tableIdx];
  if (!table) return;
  
  const filename = (currentFile.replace('.pdf','') + '_tabel' + (tableIdx+1));
  
  const r = await fetch('/api/export', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({data: table.data, format: fmt, filename})
  });
  
  if (!r.ok) {
    const d = await r.json();
    return alert('Export gagal: ' + (d.error || 'Unknown error'));
  }
  
  const blob = await r.blob();
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url;
  a.download = filename + (fmt === 'excel' ? '.xlsx' : fmt === 'json' ? '.json' : '.csv');
  a.click();
  URL.revokeObjectURL(url);
}

function showStatus(el, type, msg) {
  el.className = 'status-bar ' + type;
  el.textContent = msg;
}

function escHtml(str) {
  if (!str) return '';
  return String(str).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

// ── Init ─────────────────────────────────────────────────────────────────────
loadFileList();
</script>
</body>
</html>
'''

if __name__ == '__main__':
    print("🚀 PDF Table Extractor Dashboard")
    print("📌 Buka browser: http://localhost:5050")
    app.run(debug=False, host='0.0.0.0', port=5050)
